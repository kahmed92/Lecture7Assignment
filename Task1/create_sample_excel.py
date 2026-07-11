import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_transactions():
    """Create a detailed 3-month sample transaction history"""

    transactions = []

    # Start date: 3 months ago
    start_date = datetime.now() - timedelta(days=90)

    # Monthly salary (fixed income)
    salary_dates = [
        start_date.replace(day=1),
        (start_date + timedelta(days=30)).replace(day=1),
        (start_date + timedelta(days=60)).replace(day=1),
    ]

    for salary_date in salary_dates:
        transactions.append({
            'Date': salary_date,
            'Day': salary_date.strftime('%A'),
            'Description': 'Salary - Monthly',
            'Category': 'Income',
            'Type': 'Credit',
            'Amount': 65000,
            'Balance': 0  # Will calculate later
        })

    # PERSONAL EXPENSES
    personal_expenses = [
        # Groceries (weekly)
        {'desc': 'Groceries - Weekly Shopping', 'amount': 2500, 'frequency': 'weekly'},
        {'desc': 'Vegetables & Fruits', 'amount': 600, 'frequency': 'weekly'},
        {'desc': 'Pharmacy - Medicine', 'amount': 450, 'frequency': 'occasional'},
        {'desc': 'Clothing - T-Shirt', 'amount': 1200, 'frequency': 'occasional'},
        {'desc': 'Clothing - Jeans', 'amount': 2500, 'frequency': 'occasional'},
        {'desc': 'Haircut', 'amount': 300, 'frequency': 'monthly'},
        {'desc': 'Personal Care Products', 'amount': 800, 'frequency': 'monthly'},
    ]

    # SOCIAL EXPENSES
    social_expenses = [
        {'desc': 'Coffee Shop', 'amount': 250, 'frequency': '2-3 times/week'},
        {'desc': 'Lunch with Friends', 'amount': 600, 'frequency': 'weekly'},
        {'desc': 'Restaurant Dinner', 'amount': 1500, 'frequency': 'bi-weekly'},
        {'desc': 'Movie Tickets', 'amount': 400, 'frequency': '2-3 times/month'},
        {'desc': 'Concert Tickets', 'amount': 2500, 'frequency': 'occasional'},
        {'desc': 'Cafe - Snacks', 'amount': 300, 'frequency': '2-3 times/week'},
    ]

    # OFFICIAL EXPENSES
    official_expenses = [
        {'desc': 'Rent - Monthly', 'amount': 15000, 'day': 1, 'frequency': 'monthly'},
        {'desc': 'Internet Bill', 'amount': 1500, 'day': 5, 'frequency': 'monthly'},
        {'desc': 'Electricity Bill', 'amount': 2000, 'day': 10, 'frequency': 'monthly'},
        {'desc': 'Mobile Phone Bill', 'amount': 1200, 'day': 15, 'frequency': 'monthly'},
        {'desc': 'Insurance - Health', 'amount': 3500, 'day': 20, 'frequency': 'monthly'},
        {'desc': 'Office Supplies', 'amount': 800, 'frequency': 'occasional'},
        {'desc': 'Software Subscription', 'amount': 499, 'day': 1, 'frequency': 'monthly'},
        {'desc': 'Transport - Fuel', 'amount': 2000, 'frequency': 'weekly'},
    ]

    # Generate transaction dates
    current_date = start_date
    end_date = datetime.now()

    while current_date <= end_date:
        month_day = current_date.day
        weekday = current_date.weekday()  # 0=Monday, 6=Sunday

        # OFFICIAL EXPENSES (Fixed dates)
        for expense in official_expenses:
            if 'day' in expense and month_day == expense['day']:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': expense['desc'],
                    'Category': 'Official',
                    'Type': 'Debit',
                    'Amount': expense['amount'],
                    'Balance': 0
                })

        # PERSONAL EXPENSES (Various frequencies)
        if weekday < 5:  # Weekday
            # Groceries on Saturdays
            if weekday == 5:  # Saturday
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Groceries - Weekly Shopping',
                    'Category': 'Personal',
                    'Type': 'Debit',
                    'Amount': 2500,
                    'Balance': 0
                })
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Vegetables & Fruits',
                    'Category': 'Personal',
                    'Type': 'Debit',
                    'Amount': 600,
                    'Balance': 0
                })

            # Occasional personal expenses
            if current_date.day % 7 == 3:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Pharmacy - Medicine',
                    'Category': 'Personal',
                    'Type': 'Debit',
                    'Amount': 450,
                    'Balance': 0
                })

            if current_date.day % 15 == 0:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Personal Care Products',
                    'Category': 'Personal',
                    'Type': 'Debit',
                    'Amount': 800,
                    'Balance': 0
                })

        # SOCIAL EXPENSES (Various frequencies)
        if weekday < 5:  # Weekdays
            # Coffee 2-3 times a week
            if weekday in [1, 3]:  # Tuesday, Thursday
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Coffee Shop',
                    'Category': 'Social',
                    'Type': 'Debit',
                    'Amount': 250,
                    'Balance': 0
                })

            # Lunch with friends (weekly)
            if weekday == 2:  # Wednesday
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Lunch with Friends',
                    'Category': 'Social',
                    'Type': 'Debit',
                    'Amount': 600,
                    'Balance': 0
                })

        # Weekend social activities
        if weekday == 5:  # Saturday
            # Restaurant dinner every 2 weeks
            if current_date.day % 14 == 0 or current_date.day % 14 == 7:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Restaurant Dinner',
                    'Category': 'Social',
                    'Type': 'Debit',
                    'Amount': 1500,
                    'Balance': 0
                })

        if weekday == 4:  # Friday
            # Movie or cafe visit
            if current_date.day % 10 == 0:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Movie Tickets',
                    'Category': 'Social',
                    'Type': 'Debit',
                    'Amount': 400,
                    'Balance': 0
                })

            # Cafe snacks
            if current_date.day % 5 == 0:
                transactions.append({
                    'Date': current_date,
                    'Day': current_date.strftime('%A'),
                    'Description': 'Cafe - Snacks',
                    'Category': 'Social',
                    'Type': 'Debit',
                    'Amount': 300,
                    'Balance': 0
                })

        # Transport (weekly)
        if weekday == 6:  # Sunday
            transactions.append({
                'Date': current_date,
                'Day': current_date.strftime('%A'),
                'Description': 'Transport - Fuel',
                'Category': 'Official',
                'Type': 'Debit',
                'Amount': 2000,
                'Balance': 0
            })

        current_date += timedelta(days=1)

    # Create DataFrame
    df = pd.DataFrame(transactions)
    df = df.sort_values('Date').reset_index(drop=True)

    # Calculate running balance
    balance = 0
    balances = []
    for idx, row in df.iterrows():
        if row['Type'] == 'Credit':
            balance += row['Amount']
        else:
            balance -= row['Amount']
        balances.append(balance)

    df['Balance'] = balances

    return df


def format_excel_file(excel_file):
    """Apply formatting to Excel file"""

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Define colors
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    income_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    expense_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    personal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    social_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    official_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Apply category colors
    for row_idx in range(2, ws.max_row + 1):
        category = ws[f'D{row_idx}'].value
        trans_type = ws[f'E{row_idx}'].value

        if trans_type == 'Credit':
            for cell in ws[row_idx:row_idx]:
                for col_cell in [ws[f'{col}{row_idx}'] for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']]:
                    col_cell.fill = income_fill
        elif category == 'Personal':
            for col_cell in [ws[f'{col}{row_idx}'] for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']]:
                col_cell.fill = personal_fill
        elif category == 'Social':
            for col_cell in [ws[f'{col}{row_idx}'] for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']]:
                col_cell.fill = social_fill
        elif category == 'Official':
            for col_cell in [ws[f'{col}{row_idx}'] for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']]:
                col_cell.fill = official_fill

    # Format currency columns
    for row_idx in range(2, ws.max_row + 1):
        ws[f'F{row_idx}'].number_format = '₹ #,##0.00'
        ws[f'G{row_idx}'].number_format = '₹ #,##0.00'

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(excel_file)
    print(f"✓ Excel file formatted: {excel_file}")


def main():
    print("📊 Creating Sample Transaction History...")
    print("=" * 70)

    # Generate transactions
    print("\n1. Generating transaction data...")
    df = create_sample_transactions()
    print(f"   ✓ Generated {len(df)} transactions")
    print(f"   ✓ Date range: {df['Date'].min().date()} to {df['Date'].max().date()}")

    # Save to Excel
    excel_file = 'my_transaction_history.xlsx'
    print(f"\n2. Saving to Excel...")
    df.to_excel(excel_file, index=False, sheet_name='Transactions')
    print(f"   ✓ File created: {excel_file}")

    # Apply formatting
    print(f"\n3. Applying formatting...")
    format_excel_file(excel_file)

    # Print summary statistics
    print(f"\n4. Summary Statistics:")
    print(f"   {'=' * 70}")

    credit = df[df['Type'] == 'Credit']['Amount'].sum()
    debit = df[df['Type'] == 'Debit']['Amount'].sum()

    print(f"   Total Income (Credit):        ₹{credit:>12,.2f}")
    print(f"   Total Expenses (Debit):       ₹{debit:>12,.2f}")
    print(f"   Net Income:                   ₹{credit - debit:>12,.2f}")
    print(f"   {'=' * 70}")

    print(f"\n   Expenses by Category:")
    for category in ['Personal', 'Social', 'Official']:
        amount = df[(df['Category'] == category) & (df['Type'] == 'Debit')]['Amount'].sum()
        count = len(df[(df['Category'] == category) & (df['Type'] == 'Debit')])
        pct = (amount / debit * 100) if debit > 0 else 0
        print(f"   • {category:<15} ₹{amount:>10,.2f} ({count:>3} transactions, {pct:>5.1f}%)")

    print(f"\n   Transactions by Day:")
    day_counts = df['Day'].value_counts().sort_values(ascending=False)
    for day, count in day_counts.items():
        print(f"   • {day:<12} {count:>3} transactions")

    print(f"\n{'=' * 70}")
    print(f"✅ SUCCESS: Your transaction history is ready!")
    print(f"   File: {excel_file}")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    main()

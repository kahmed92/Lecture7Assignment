#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create workbook for Expected Payments
wb_expected = openpyxl.Workbook()
ws_expected = wb_expected.active
ws_expected.title = "Expected Payments"

# Add headers
headers_expected = ["Name", "Amount", "Due Date", "Description"]
for col, header in enumerate(headers_expected, 1):
    cell = ws_expected.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Sample expected payments data
expected_data = [
    ["John Smith", 500, "2026-07-01", "Q3 Sponsorship"],
    ["Sarah Johnson", 1000, "2026-07-02", "Event Ticket Sales"],
    ["Michael Chen", 750, "2026-07-03", "Workshop Fee"],
    ["Emma Davis", 250, "2026-07-04", "Membership Renewal"],
    ["Robert Wilson", 600, "2026-07-05", "Donation"],
    ["Lisa Anderson", 450, "2026-07-06", "Course Registration"],
    ["James Martinez", 800, "2026-07-07", "Conference Attendance"],
    ["Jennifer Taylor", 350, "2026-07-08", "Raffle Ticket"],
    ["David Brown", 1100, "2026-07-09", "Sponsorship Package"],
    ["Patricia Lee", 525, "2026-07-10", "Workshop Fee"],
]

total_expected = 0
for row_idx, row_data in enumerate(expected_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_expected.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left" if col_idx <= 1 else "center")
    total_expected += row_data[1]

# Add total row
total_row = len(expected_data) + 3
ws_expected.cell(row=total_row, column=1).value = "TOTAL EXPECTED"
ws_expected.cell(row=total_row, column=1).font = Font(bold=True)
ws_expected.cell(row=total_row, column=2).value = total_expected
ws_expected.cell(row=total_row, column=2).font = Font(bold=True)
ws_expected.cell(row=total_row, column=2).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

# Adjust column widths
ws_expected.column_dimensions['A'].width = 20
ws_expected.column_dimensions['B'].width = 12
ws_expected.column_dimensions['C'].width = 15
ws_expected.column_dimensions['D'].width = 25

wb_expected.save("expected_payments.xlsx")
print(f"✓ Created expected_payments.xlsx")
print(f"  Total Expected: ${total_expected}")

# Create workbook for Actual Payment Records
wb_actual = openpyxl.Workbook()
ws_actual = wb_actual.active
ws_actual.title = "Payment Records"

# Add headers
headers_actual = ["Date", "Payer Name", "Amount", "Payment Method", "Reference/Notes"]
for col, header in enumerate(headers_actual, 1):
    cell = ws_actual.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Sample actual payment records with intentional discrepancies
payment_records = [
    ["2026-07-01", "John Smith", 500, "Bank Transfer", "Q3 Sponsorship"],
    ["2026-07-02", "Sarah Johnson", 1000, "Credit Card", "Event Ticket Sales"],
    ["2026-07-03", "M. Chen", 750, "Check", "Workshop - Michael Chen"],
    ["2026-07-04", "E. Davis", 250, "PayPal", "Membership - Emma Davis"],
    ["2026-07-05", "Robert Wilson", 600, "Bank Transfer", "Donation"],
    ["2026-07-06", "Lisa Anderson", 450, "Credit Card", "Course Registration"],
    ["2026-07-07", "James Martinez", 800, "Check", "Conference Attendance"],
    ["2026-07-08", "Jennifer Taylor", 350, "PayPal", "Raffle Ticket"],
    ["2026-07-09", "David Brown", 1100, "Bank Transfer", "Sponsorship Package"],
    ["2026-07-10", "Patricia Lee", 525, "Credit Card", "Workshop Fee"],
    ["2026-07-10", "Patricia Lee", 100, "Check", "DUPLICATE ENTRY - Error"],
    ["2026-07-11", "Unknown Donor", 250, "Cash", "Anonymous Donation"],
    ["2026-07-12", "Sarah Johnson", 500, "Bank Transfer", "Additional Event Sponsorship"],
    ["2026-07-12", "Robert Wilson", 100, "Credit Card", "Late Fee"],
]

total_actual = 0
for row_idx, row_data in enumerate(payment_records, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_actual.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left" if col_idx in [1, 2, 5] else "center")
    total_actual += row_data[2]

# Add total row
total_row = len(payment_records) + 3
ws_actual.cell(row=total_row, column=1).value = "TOTAL RECORDED"
ws_actual.cell(row=total_row, column=1).font = Font(bold=True)
ws_actual.cell(row=total_row, column=3).value = total_actual
ws_actual.cell(row=total_row, column=3).font = Font(bold=True)
ws_actual.cell(row=total_row, column=3).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

# Adjust column widths
ws_actual.column_dimensions['A'].width = 15
ws_actual.column_dimensions['B'].width = 20
ws_actual.column_dimensions['C'].width = 12
ws_actual.column_dimensions['D'].width = 18
ws_actual.column_dimensions['E'].width = 30

wb_actual.save("payment_records.xlsx")
print(f"✓ Created payment_records.xlsx")
print(f"  Total Recorded: ${total_actual}")
print(f"\nDiscrepancy: ${total_actual - total_expected} (Difference of ${abs(total_actual - total_expected)})")
